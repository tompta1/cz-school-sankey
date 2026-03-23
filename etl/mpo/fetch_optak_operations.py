#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import urljoin

from openpyxl import load_workbook

from _common import RAW_ROOT, USER_AGENT, fetch_bytes, sha256_bytes, timestamp_label, write_sidecar

DATASET_CODE = "mpo_optak_operations"
LISTING_URL = "https://www.dotaceeu.cz/cs/statistiky-a-analyzy/seznam-operaci-%28prijemcu%29"
OUTPUT_FILE_NAME = "optak-operations-normalized.csv"


@dataclass(frozen=True)
class OptakRow:
    reporting_year: int
    region_code: str
    region_name: str
    recipient_key: str
    recipient_name: str
    recipient_ico: str
    project_id: str
    project_name: str
    priority_name: str
    specific_objective_name: str
    intervention_name: str
    allocated_total_czk: Decimal
    union_support_czk: Decimal
    national_public_czk: Decimal
    charged_total_czk: Decimal
    source_url: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch MPO OP TAK operations from DotaceEU")
    parser.add_argument("--year", action="append", type=int, required=True, help="Workbook year to fetch. Can be used multiple times.")
    parser.add_argument("--snapshot", default=None, help="Snapshot label, defaults to YYYYMMDD")
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=RAW_ROOT / DATASET_CODE,
        help="Output directory",
    )
    return parser.parse_args()


def normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def normalize_ico(value: object | None) -> str:
    digits = "".join(ch for ch in normalize_text(value) if ch.isdigit())
    if not digits:
        return ""
    return digits.zfill(8) if len(digits) <= 8 else digits


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def parse_decimal(value: object | None) -> Decimal:
    text = normalize_text(value)
    if not text:
        return Decimal("0")
    text = text.replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def find_workbook_url(year: int) -> str:
    html = fetch_bytes(LISTING_URL).decode("utf-8", "ignore")
    matches = re.findall(r'href="([^"]+Seznam-operaci[^"]+_21\.xlsx(?:\.aspx\?ext=\.xlsx)?)"', html, flags=re.IGNORECASE)
    exact = [urljoin(LISTING_URL, href) for href in matches if f"/{year}_12_" in href]
    if exact:
        return exact[0]
    fallback = [urljoin(LISTING_URL, href) for href in matches if f"/{year}_" in href]
    if fallback:
        return sorted(fallback)[-1]
    raise RuntimeError(f"No DotaceEU 21+ workbook link found for year {year}")


def recipient_key(recipient_ico: str, recipient_name: str, project_id: str) -> str:
    if recipient_ico:
        return recipient_ico
    return f"JI:{normalize_key(recipient_name)}|{project_id}"


def parse_workbook_rows(payload: bytes, year: int, source_url: str) -> list[OptakRow]:
    workbook = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(min_row=4, values_only=True)
    headers = [normalize_text(value) for value in next(rows)]
    next(rows, None)
    column_index = {name: index for index, name in enumerate(headers)}

    def value(row: tuple[object, ...], column_name: str) -> object | None:
        index = column_index.get(column_name)
        if index is None or index >= len(row):
            return None
        return row[index]

    parsed: list[OptakRow] = []
    for row in sheet.iter_rows(min_row=6, values_only=True):
        program_code = normalize_text(value(row, "Číslo programu"))
        if program_code != "01":
            continue

        project_id = normalize_text(value(row, "Registrační číslo projektu"))
        recipient_name = normalize_text(value(row, "Příjemce - název"))
        project_name = normalize_text(value(row, "Název projektu"))
        if not project_id or not recipient_name or not project_name:
          continue

        allocated_total_czk = parse_decimal(value(row, "Finanční prostředky v právních aktech celkové způsobilé výdaje CZK"))
        if allocated_total_czk <= 0:
            continue

        recipient_ico = normalize_ico(value(row, "IČ příjemce"))
        region_code = normalize_text(value(row, "Místo realizace Kód NUTS 3"))
        region_name = normalize_text(value(row, "Místo realizace Název NUTS 3"))
        priority_name = normalize_text(value(row, "Název priority"))
        specific_objective_name = normalize_text(value(row, "Název specifického cíle"))
        intervention_name = normalize_text(value(row, "Oblast intervence - název"))
        union_support_czk = parse_decimal(value(row, "Finanční prostředky v právních aktech příspěvek Unie CZK"))
        national_public_czk = parse_decimal(value(row, "Finanční prostředky v právních aktech národní veřejné zdroje CZK"))
        charged_total_czk = parse_decimal(value(row, "Finanční prostředky vyúčtované v žádostech o platbu celkové způsobilé výdaje CZK"))

        parsed.append(
            OptakRow(
                reporting_year=year,
                region_code=region_code,
                region_name=region_name,
                recipient_key=recipient_key(recipient_ico, recipient_name, project_id),
                recipient_name=recipient_name,
                recipient_ico=recipient_ico,
                project_id=project_id,
                project_name=project_name,
                priority_name=priority_name,
                specific_objective_name=specific_objective_name,
                intervention_name=intervention_name,
                allocated_total_czk=allocated_total_czk,
                union_support_czk=union_support_czk,
                national_public_czk=national_public_czk,
                charged_total_czk=charged_total_czk,
                source_url=source_url,
            )
        )

    return parsed


def write_snapshot(
    *,
    out_dir: Path,
    snapshot: str,
    rows: list[OptakRow],
    metadata_rows: list[dict[str, object]],
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    data_path = out_dir / f"{snapshot}__{OUTPUT_FILE_NAME}"
    fieldnames = [
        "reporting_year",
        "region_code",
        "region_name",
        "recipient_key",
        "recipient_name",
        "recipient_ico",
        "project_id",
        "project_name",
        "priority_name",
        "specific_objective_name",
        "intervention_name",
        "allocated_total_czk",
        "union_support_czk",
        "national_public_czk",
        "charged_total_czk",
        "source_url",
    ]

    with data_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "reporting_year": row.reporting_year,
                    "region_code": row.region_code,
                    "region_name": row.region_name,
                    "recipient_key": row.recipient_key,
                    "recipient_name": row.recipient_name,
                    "recipient_ico": row.recipient_ico,
                    "project_id": row.project_id,
                    "project_name": row.project_name,
                    "priority_name": row.priority_name,
                    "specific_objective_name": row.specific_objective_name,
                    "intervention_name": row.intervention_name,
                    "allocated_total_czk": f"{row.allocated_total_czk:.2f}",
                    "union_support_czk": f"{row.union_support_czk:.2f}",
                    "national_public_czk": f"{row.national_public_czk:.2f}",
                    "charged_total_czk": f"{row.charged_total_czk:.2f}",
                    "source_url": row.source_url,
                }
            )

    write_sidecar(
        data_path,
        {
            "dataset_code": DATASET_CODE,
            "downloaded_at": datetime.now(UTC).isoformat(),
            "row_count": len(rows),
            "year_count": len(metadata_rows),
            "years": sorted({int(row["year"]) for row in metadata_rows}),
            "sources": metadata_rows,
            "generator": "etl/mpo/fetch_optak_operations.py",
            "listing_url": LISTING_URL,
            "user_agent": USER_AGENT,
        },
    )
    return data_path


def main() -> None:
    args = parse_args()
    years = sorted(set(args.year))
    snapshot = timestamp_label(args.snapshot)

    rows: list[OptakRow] = []
    metadata_rows: list[dict[str, object]] = []
    for year in years:
        workbook_url = find_workbook_url(year)
        payload = fetch_bytes(workbook_url)
        year_rows = parse_workbook_rows(payload, year, workbook_url)
        rows.extend(year_rows)
        metadata_rows.append(
            {
                "year": year,
                "source_url": workbook_url,
                "content_sha256": sha256_bytes(payload),
                "row_count": len(year_rows),
            }
        )

    rows.sort(key=lambda row: (row.reporting_year, -row.allocated_total_czk, row.recipient_name, row.project_id))
    data_path = write_snapshot(
        out_dir=args.out_dir,
        snapshot=snapshot,
        rows=rows,
        metadata_rows=metadata_rows,
    )

    print(f"Wrote {data_path}")
    print(f"Years: {', '.join(str(year) for year in years)}")
    print(f"Rows written: {len(rows)}")


if __name__ == "__main__":
    main()
