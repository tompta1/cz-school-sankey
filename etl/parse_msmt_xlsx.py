#!/usr/bin/env python3
"""Parse the official MŠMT 'Podrobný rozpis rozpočtu' XLSX into the ETL CSVs.

Usage:
    python3 etl/parse_msmt_xlsx.py --year 2025 [--kraj Středočeský] [--top 30]

Outputs:
    etl/data/raw/<year>/school_entities.csv
    etl/data/raw/<year>/msmt_allocations.csv

Column mapping (row 8 of the XLSX is the machine-readable header):
    ICO        – school legal entity IČO  (join key)
    ICO_ZRIZ   – founder IČO
    ZRIZ       – founder type code (2 = kraj/obec ÚSC, 7 = Praha)
    KRAJ       – region name
    NIV_CELKEM – total MŠMT allocation (the inflow edge value)
    PLATY_CELKEM – total salaries (ped + nonped)
    ODVODY_CELKEM – insurance contributions
    FKSP_CELKEM – FKSP contributions
    ONIV_CELKEM – other non-investment expenditures

Pedagogical salary columns (PP = pedagogický pracovník, AP = asistent pedagoga):
    TARIF_PP_*/NADTARIF_PP_*/AO_* – tariff/above-tariff/adaptation for PP per school type
    TARIF_PPA_*/NADTARIF_PPA_*    – same for AP (asistent pedagoga)
    PLAT_PP_* (VOŠ, ZUŠ)         – flat salary column for those types

Non-pedagogical salary columns (NPZ = nepedagogický zaměstnanec):
    PLAT_NPZ_R, PLAT_NPZ_P       – directorate and other workplaces
    PLAT_NPZ_MS/ZS/SS/KN/VOS/ZUS – per school type
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW_ROOT = ROOT / "etl" / "data" / "raw"

# IČOs of the 13 kraj + Praha (used to distinguish kraj vs obec founders)
KRAJ_ICOS: frozenset[str] = frozenset({
    "00064581",  # Hlavní město Praha
    "70891095",  # Středočeský kraj
    "70890650",  # Jihočeský kraj
    "70890366",  # Plzeňský kraj
    "70891168",  # Karlovarský kraj
    "70892156",  # Ústecký kraj
    "70891508",  # Liberecký kraj
    "70889546",  # Královéhradecký kraj
    "71004741",  # Pardubický kraj
    "60609460",  # Olomoucký kraj
    "70890692",  # Moravskoslezský kraj
    "70888337",  # Jihomoravský kraj
    "70891320",  # Zlínský kraj
    "70890749",  # Kraj Vysočina
})

KRAJ_NAME_TO_ICO: dict[str, str] = {
    "Praha":           "00064581",
    "Středočeský":     "70891095",
    "Jihočeský":       "70890650",
    "Plzeňský":        "70890366",
    "Karlovarský":     "70891168",
    "Ústecký":         "70892156",
    "Liberecký":       "70891508",
    "Královéhradecký": "70889546",
    "Pardubický":      "71004741",
    "Olomoucký":       "60609460",
    "Moravskoslezský": "70890692",
    "Jihomoravský":    "70888337",
    "Zlínský":         "70891320",
    "Vysočina":        "70890749",
}

KRAJ_NAMES: dict[str, str] = {
    "00064581": "Hlavní město Praha",
    "70891095": "Středočeský kraj",
    "70890650": "Jihočeský kraj",
    "70890366": "Plzeňský kraj",
    "70891168": "Karlovarský kraj",
    "70892156": "Ústecký kraj",
    "70891508": "Liberecký kraj",
    "70889546": "Královéhradecký kraj",
    "71004741": "Pardubický kraj",
    "60609460": "Olomoucký kraj",
    "70890692": "Moravskoslezský kraj",
    "70888337": "Jihomoravský kraj",
    "70891320": "Zlínský kraj",
    "70890749": "Kraj Vysočina",
}

# Pedagogical salary column indices (0-based, from machine-readable header row)
PP_COLS = [16, 17, 18, 21, 22,   # MŠ: TARIF_PP, NADTARIF_PP, AO, TARIF_PPA, NADTARIF_PPA
           28, 29, 30, 33, 34,   # ZŠ
           40, 41, 44, 45,       # ŠD
           47, 48, 49, 52, 53,   # SŠ
           59, 60, 61,           # KN
           68,                   # VOŠ: PLAT_PP_VOS
           73, 74]               # ZUŠ: PLAT_PP_ZUS, AO_ZUS

# Non-pedagogical salary column indices
NPZ_COLS = [12, 14,             # ředitelství + další pracoviště
            24, 36, 55, 64, 70, 76]  # MŠ, ZŠ, SŠ, KN, VOŠ, ZUŠ


def safe_int(value: object) -> int:
    if value is None:
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def normalize_ico(raw: object) -> str:
    """Return an 8-digit zero-padded IČO string."""
    return str(int(raw)).zfill(8) if raw is not None else ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse MŠMT XLSX into ETL CSVs")
    parser.add_argument("--year", type=int, default=2025)
    parser.add_argument("--kraj", default=None,
                        help="Filter to one region name (e.g. 'Středočeský'). "
                             "Omit to include all regions.")
    parser.add_argument("--top", type=int, default=50,
                        help="Keep only the top N schools by NIV_CELKEM (default 50). "
                             "Use 0 for all.")
    parser.add_argument("--xlsx", default=None,
                        help="Path to the MŠMT XLSX file. "
                             "Defaults to etl/data/raw/<year>/msmt_<year>_raw.xlsx")
    parser.add_argument("--founder-lookup", default=None,
                        help="Path to a school_entities.csv from another year to use as "
                             "founder fallback when ICO_ZRIZ column is absent (e.g. 2024 XLSX).")
    return parser.parse_args()


def load_founder_lookup(csv_path: str) -> dict[str, dict]:
    """Load ico -> {founder_ico, founder_name, founder_type} from a school_entities.csv."""
    lookup: dict[str, dict] = {}
    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            ico = row.get("ico", "").strip()
            founder_id = row.get("founder_id", "").strip()
            founder_ico = founder_id.removeprefix("founder:").strip()
            if ico and founder_ico:
                lookup[ico] = {
                    "founder_ico": founder_ico,
                    "founder_name": row.get("founder_name", ""),
                    "founder_type": row.get("founder_type", "obec"),
                }
    return lookup


def load_rows(xlsx_path: Path, year: int) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheet_name = f"Podrobný rozpis rozpočtu {year}"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # Row 8 (1-based) is the machine-readable header; data starts at row 9
    raw = list(ws.iter_rows(min_row=8, values_only=True))
    header = raw[0]
    data_rows = raw[1:]

    records: list[dict[str, object]] = []
    for row in data_rows:
        if not row[2]:  # skip rows without ICO
            continue
        record = {header[i]: row[i] for i in range(len(header)) if header[i]}
        records.append(record)

    wb.close()
    return records


def build_csvs(records: list[dict], year: int, kraj_filter: str | None, top_n: int, founder_lookup: dict[str, dict] | None = None) -> tuple[list[dict], list[dict]]:
    if kraj_filter:
        records = [r for r in records if (r.get("KRAJ") or "").startswith(kraj_filter)]

    # Sort by NIV_CELKEM descending and keep top N
    records.sort(key=lambda r: safe_int(r.get("NIV_CELKEM")), reverse=True)
    if top_n > 0:
        records = records[:top_n]

    entities: list[dict] = []
    allocations: list[dict] = []

    for r in records:
        ico = normalize_ico(r.get("ICO"))
        ico_zriz = normalize_ico(r.get("ICO_ZRIZ"))
        if not ico:
            continue

        # 2025+ XLSX has ICO_ZRIZ; 2024 and earlier only have a ZRIZ type code
        # (2 = obec/kraj ÚSC, 7 = kraj-funded school — not just Praha).
        # Resolve via external lookup first, then KRAJ-name fallback for ZRIZ=7.
        if not ico_zriz:
            zriz_code = r.get("ZRIZ")
            if founder_lookup and ico in founder_lookup:
                ico_zriz = founder_lookup[ico]["founder_ico"]
            elif zriz_code == 7:
                # Kraj-funded but not in lookup: derive founder from region name
                kraj_name = (r.get("KRAJ") or "").strip()
                ico_zriz = KRAJ_NAME_TO_ICO.get(kraj_name, "")
        founder_type = "kraj" if ico_zriz in KRAJ_ICOS else "obec"
        founder_name = KRAJ_NAMES.get(ico_zriz) or f"Zřizovatel IČO {ico_zriz}"
        founder_id = f"founder:{ico_zriz}"
        kraj = r.get("KRAJ") or ""

        institution_id = f"school:{ico}"
        institution_name = f"IČO {ico}"  # placeholder; replace with registry names

        entities.append({
            "institution_id": institution_id,
            "institution_name": institution_name,
            "ico": ico,
            "founder_id": founder_id,
            "founder_name": founder_name,
            "founder_type": founder_type,
            "municipality": "",
            "region": kraj,
        })

        niv = safe_int(r.get("NIV_CELKEM"))
        ped = sum(safe_int(r.get(k)) for k in _cols_by_index(r, PP_COLS))
        nonped_keys = [k for k in NPZ_COLS_NAMES if k in r]
        nonped = sum(safe_int(r.get(k)) for k in nonped_keys)
        oniv = safe_int(r.get("ONIV_CELKEM"))
        odvody = safe_int(r.get("ODVODY_CELKEM"))
        fksp = safe_int(r.get("FKSP_CELKEM"))
        other = odvody + fksp
        # Sanity: ped + nonped + oniv + other should ≈ niv
        # (small rounding differences from the Excel are expected)

        allocations.append({
            "institution_id": institution_id,
            "pedagogical_amount": ped,
            "nonpedagogical_amount": nonped,
            "oniv_amount": oniv,
            "other_amount": other,
            "operations_amount": 0,
            "investment_amount": 0,
            "bucket_basis": "allocated",
            "bucket_certainty": "observed",
            "_niv_total": niv,  # for cross-check, not written to CSV
        })

    return entities, allocations


# Column name lists for PP and NPZ (derived from the machine-readable header)
PP_COL_NAMES = [
    "TARIF_PP_MS", "NADTARIF_PP_MS", "AO_MS", "TARIF_PPA_MS", "NADTARIF_PPA_MS",
    "TARIF_PP_ZS", "NADTARIF_PP_ZS", "AO_ZS", "TARIF_PPA_ZS", "NADTARIF_PPA_ZS",
    "TARIF_PP_SD", "NADTARIF_PP_SD", "TARIF_PPA_SD", "NADTARIF_PPA_SD",
    "TARIF_PP_SS", "NADTARIF_PP_SS", "AO_SS", "TARIF_PPA_SS", "NADTARIF_PPA_SS",
    "TARIF_PP_KN", "NADTARIF_PP_KN", "AO_KN",
    "PLAT_PP_VOS",
    "PLAT_PP_ZUS", "AO_ZUS",
]

NPZ_COLS_NAMES = [
    "PLAT_NPZ_R", "PLAT_NPZ_P",
    "PLAT_NPZ_MS", "PLAT_NPZ_ZS", "PLAT_NPZ_SS", "PLAT_NPZ_KN", "PLAT_NPZ_VOS", "PLAT_NPZ_ZUS",
]


def _cols_by_index(record: dict, _unused: list) -> list[str]:
    """Return PP column names that exist in the record (handles missing school types)."""
    return [k for k in PP_COL_NAMES if k in record]


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    year = args.year

    xlsx_path = Path(args.xlsx) if args.xlsx else RAW_ROOT / str(year) / f"msmt_{year}_raw.xlsx"
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    print(f"Loading {xlsx_path} …")
    records = load_rows(xlsx_path, year)
    print(f"  {len(records)} school entities loaded")

    founder_lookup = load_founder_lookup(args.founder_lookup) if args.founder_lookup else None
    if founder_lookup:
        print(f"  Loaded founder lookup with {len(founder_lookup)} entries from {args.founder_lookup}")
    entities, allocations = build_csvs(records, year, args.kraj, args.top, founder_lookup)
    print(f"  After filter/top: {len(entities)} schools")

    out_dir = RAW_ROOT / str(year)

    entity_fields = ["institution_id", "institution_name", "ico", "founder_id",
                     "founder_name", "founder_type", "municipality", "region"]
    write_csv(out_dir / "school_entities.csv", entities, entity_fields)
    print(f"  Wrote school_entities.csv ({len(entities)} rows)")

    alloc_fields = ["institution_id", "pedagogical_amount", "nonpedagogical_amount",
                    "oniv_amount", "other_amount", "operations_amount", "investment_amount",
                    "bucket_basis", "bucket_certainty"]
    write_csv(out_dir / "msmt_allocations.csv", allocations, alloc_fields)
    print(f"  Wrote msmt_allocations.csv ({len(allocations)} rows)")

    total_niv = sum(a["_niv_total"] for a in allocations)
    total_ped = sum(a["pedagogical_amount"] for a in allocations)
    total_nonped = sum(a["nonpedagogical_amount"] for a in allocations)
    total_oniv = sum(a["oniv_amount"] for a in allocations)
    total_other = sum(a["other_amount"] for a in allocations)
    reconstructed = total_ped + total_nonped + total_oniv + total_other
    print(f"\nCross-check (should be close):")
    print(f"  NIV_CELKEM sum:          {total_niv:>15,} CZK")
    print(f"  Ped+Nonped+ONIV+Other:   {reconstructed:>15,} CZK")
    print(f"  Difference:              {total_niv - reconstructed:>15,} CZK")


if __name__ == "__main__":
    main()
