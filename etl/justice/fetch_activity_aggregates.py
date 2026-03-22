#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from _common import RAW_ROOT, USER_AGENT, fetch_bytes, sha256_bytes, timestamp_label

DATASET_CODE = "justice_activity_aggregates"
OUTPUT_FILE_NAME = "justice-activity-aggregates.csv"

COURT_SOURCE_URL = "https://msp.gov.cz/documents/d/msp/data_soudy_2024-xlsm"
PRISON_SOURCE_URL = (
    "https://www.vscr.cz/media/organizacni-jednotky/generalni-reditelstvi/odbor-spravni/"
    "statistiky/rocenky/statisticka-rocenka-vezenske-sluzby-ceske-republiky-za-rok-2024.pdf"
)

DECIMAL_RE = re.compile(r"\d[\d ]*,\d+")

COURT_SHEETS = {
    "courts_district_disposed_total": [
        "Přehled_trest_2024",
        "Přehled_civil_2024",
        "Přehled_opatro_2024",
    ],
    "courts_regional_disposed_total": [
        "Přehled_trest_2024_KS",
        "Přehled_To_2024_KS",
        "Přehled_Cm_2024_KS",
        "Přehled_C_2024_KS",
        "Přehled_Co_2024_KS",
        "Přehled_A_2024_KS",
        "Přehled_INS_2024_KS",
        "Přehled_ICm_2024_KS",
    ],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch justice activity denominators from official court and prison sources")
    parser.add_argument("--year", type=int, default=2024, help="Reporting year currently supported")
    parser.add_argument("--snapshot", default=None, help="Snapshot label, defaults to YYYYMMDD")
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=RAW_ROOT / DATASET_CODE,
        help="Output directory",
    )
    return parser.parse_args()


def disposed_total_for_sheet(workbook_path: Path, sheet_name: str) -> int:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    disposed_column = None
    start_row = None
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        values = [str(value).strip() if value is not None else "" for value in row]
        for column_index, value in enumerate(values):
            if "Vyřízeno*" in value:
                disposed_column = column_index
                start_row = row_index + 1
                break
        if disposed_column is not None:
            break

    if disposed_column is None or start_row is None:
        raise RuntimeError(f"Could not find disposed column in sheet {sheet_name!r}")

    total = 0
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        first = row[0]
        if not isinstance(first, (int, float)):
            continue
        value = row[disposed_column]
        if isinstance(value, (int, float)):
            total += int(round(value))
    workbook.close()
    return total


def court_activity_rows(workbook_path: Path) -> list[dict[str, object]]:
    district_total = sum(disposed_total_for_sheet(workbook_path, name) for name in COURT_SHEETS["courts_district_disposed_total"])
    regional_total = sum(disposed_total_for_sheet(workbook_path, name) for name in COURT_SHEETS["courts_regional_disposed_total"])
    total = district_total + regional_total

    return [
        {
            "reporting_year": 2024,
            "activity_domain": "courts",
            "metric_code": "courts_disposed_total",
            "metric_name": "Vyřízené věci celkem",
            "count_value": total,
            "source_url": COURT_SOURCE_URL,
        },
        {
            "reporting_year": 2024,
            "activity_domain": "courts",
            "metric_code": "courts_district_disposed_total",
            "metric_name": "Vyřízené věci okresních soudů",
            "count_value": district_total,
            "source_url": COURT_SOURCE_URL,
        },
        {
            "reporting_year": 2024,
            "activity_domain": "courts",
            "metric_code": "courts_regional_disposed_total",
            "metric_name": "Vyřízené věci krajských soudů",
            "count_value": regional_total,
            "source_url": COURT_SOURCE_URL,
        },
    ]


def prison_activity_rows(pdf_path: Path) -> list[dict[str, object]]:
    reader = PdfReader(str(pdf_path))
    page_text = (reader.pages[76].extract_text() or "").replace("\xa0", " ")
    target_line = None
    for line in page_text.splitlines():
        if line.strip().startswith("Celkem "):
            target_line = " ".join(line.split())
            break
    if target_line is None:
        raise RuntimeError("Could not find average daily prison population line in prison yearbook")

    numeric_tokens = DECIMAL_RE.findall(target_line)
    average_daily_inmates = float(numeric_tokens[5].replace(" ", "").replace(",", "."))

    return [
        {
            "reporting_year": 2024,
            "activity_domain": "prison_service",
            "metric_code": "prison_average_daily_inmates_total",
            "metric_name": "Průměrný denní stav vězněných osob",
            "count_value": average_daily_inmates,
            "source_url": PRISON_SOURCE_URL,
        },
    ]


def write_snapshot(*, out_dir: Path, snapshot: str, rows: list[dict[str, object]], metadata: dict) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    data_path = out_dir / f"{snapshot}__{OUTPUT_FILE_NAME}"
    fieldnames = [
        "reporting_year",
        "activity_domain",
        "metric_code",
        "metric_name",
        "count_value",
        "source_url",
    ]

    with data_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    sidecar_path = data_path.with_suffix(data_path.suffix + ".download.json")
    sidecar_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    return data_path


def main() -> None:
    args = parse_args()
    if args.year != 2024:
        raise SystemExit("Justice activity denominators are currently implemented for 2024 only")

    snapshot = timestamp_label(args.snapshot)
    out_dir = args.out_dir

    court_bytes = fetch_bytes(COURT_SOURCE_URL)
    court_path = Path("/tmp/justice_courts_2024.xlsm")
    court_path.write_bytes(court_bytes)

    prison_bytes = fetch_bytes(PRISON_SOURCE_URL)
    prison_path = Path("/tmp/justice_prisons_2024.pdf")
    prison_path.write_bytes(prison_bytes)

    rows = [*court_activity_rows(court_path), *prison_activity_rows(prison_path)]

    metadata = {
        "dataset_code": DATASET_CODE,
        "downloaded_at": datetime.now(UTC).isoformat(),
        "row_count": len(rows),
        "years": [args.year],
        "sources": [
            {
                "reporting_year": args.year,
                "source_url": COURT_SOURCE_URL,
                "sha256": sha256_bytes(court_bytes),
                "size_bytes": len(court_bytes),
            },
            {
                "reporting_year": args.year,
                "source_url": PRISON_SOURCE_URL,
                "sha256": sha256_bytes(prison_bytes),
                "size_bytes": len(prison_bytes),
            },
        ],
        "generator": "etl/justice/fetch_activity_aggregates.py",
        "user_agent": USER_AGENT,
    }
    data_path = write_snapshot(out_dir=out_dir, snapshot=snapshot, rows=rows, metadata=metadata)

    print(f"Wrote {data_path}")
    print(f"Rows written: {len(rows)}")


if __name__ == "__main__":
    main()
