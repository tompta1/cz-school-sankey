#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from _common import RAW_ROOT, fetch_bytes, sha256_bytes, timestamp_label, write_sidecar

DATASET_CODE = "mzv_aid_operations"
OUTPUT_FILE_NAME = "mzv-aid-operations-normalized.csv"
MZV_WORKBOOK_URL = "https://mzv.gov.cz/file/5896230/Priloha_3_MZV_resorty_projekty_2024.xlsx"
CRA_WORKBOOK_URL = "https://mzv.gov.cz/file/5896228/Priloha_2_CRA_Projekty_2024.xlsx"


@dataclass(frozen=True)
class AidRow:
    reporting_year: int
    branch_code: str
    branch_name: str
    source_workbook: str
    section_code: str | None
    section_name: str | None
    country_name: str
    sector_name: str | None
    manager_code: str | None
    manager_name: str | None
    recipient_key: str
    recipient_name: str
    recipient_ico: str | None
    project_key: str
    project_name: str
    planned_czk: Decimal
    actual_czk: Decimal
    source_url: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch MZV / ČRA development and humanitarian operations")
    parser.add_argument("--year", action="append", type=int, required=True, help="Reporting year to fetch. Only 2024 is currently supported.")
    parser.add_argument("--snapshot", default=None, help="Snapshot label, defaults to YYYYMMDD")
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=RAW_ROOT / DATASET_CODE,
        help="Output directory",
    )
    return parser.parse_args()


def normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return "-".join(text.split())


def parse_decimal(value: object | None) -> Decimal:
    text = normalize_text(value)
    if not text:
        return Decimal("0")
    text = text.replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def project_key(source_workbook: str, branch_code: str, country_name: str, recipient_name: str, project_name: str) -> str:
    return normalize_key("|".join([source_workbook, branch_code, country_name, recipient_name, project_name]))


def recipient_key(source_workbook: str, recipient_name: str) -> str:
    return normalize_key(f"{source_workbook}|{recipient_name}")


def parse_mzv_workbook(payload: bytes, year: int) -> list[AidRow]:
    workbook = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook["List1"]
    current_section = ""
    rows: list[AidRow] = []

    for row in sheet.iter_rows(values_only=True):
        values = list(row[:8])
        first = normalize_text(values[0])
        if first and all(value is None for value in values[1:]):
            current_section = first
            continue

        country_name = first
        project_name = normalize_text(values[3])
        recipient_name = normalize_text(values[4])
        if not country_name or not project_name or not recipient_name or country_name.startswith("CELKEM"):
            continue

        branch_code = "HUMANITARIAN" if "humanit" in current_section.lower() else "DEVELOPMENT"
        branch_name = "Humanitární pomoc a stabilizační projekty" if branch_code == "HUMANITARIAN" else "Rozvojová spolupráce a zahraniční projekty"
        planned_czk = parse_decimal(values[5])
        actual_czk = parse_decimal(values[6])
        reserve_czk = parse_decimal(values[7])
        actual = actual_czk if actual_czk > 0 else reserve_czk if reserve_czk > 0 else planned_czk
        if actual <= 0:
            continue

        section_match = re.match(r"^(?P<code>\d+(?:\.\d+)*)\s+(?P<label>.+)$", current_section)
        section_code = section_match.group("code") if section_match else None
        section_name = section_match.group("label") if section_match else current_section or None
        manager = normalize_text(values[2]) or "MZV"
        sector_name = normalize_text(values[1]) or None
        rows.append(
            AidRow(
                reporting_year=year,
                branch_code=branch_code,
                branch_name=branch_name,
                source_workbook="MZV",
                section_code=section_code,
                section_name=section_name,
                country_name=country_name,
                sector_name=sector_name,
                manager_code=manager,
                manager_name=manager,
                recipient_key=recipient_key("MZV", recipient_name),
                recipient_name=recipient_name,
                recipient_ico=None,
                project_key=project_key("MZV", branch_code, country_name, recipient_name, project_name),
                project_name=project_name,
                planned_czk=planned_czk,
                actual_czk=actual,
                source_url=MZV_WORKBOOK_URL,
            )
        )

    return rows


def parse_cra_workbook(payload: bytes, year: int) -> list[AidRow]:
    workbook = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook["ZRS ČR 2024"]
    rows: list[AidRow] = []

    for row in sheet.iter_rows(values_only=True):
        values = list(row[:6])
        country_name = normalize_text(values[0])
        sector_name = normalize_text(values[1]) or None
        project_name = normalize_text(values[2])
        recipient_name = normalize_text(values[3])
        if not country_name or not project_name or not recipient_name or country_name.startswith("CELKEM"):
            continue

        actual_czk = parse_decimal(values[4])
        if actual_czk <= 0:
            continue

        rows.append(
            AidRow(
                reporting_year=year,
                branch_code="DEVELOPMENT",
                branch_name="Rozvojová spolupráce a zahraniční projekty",
                source_workbook="CRA",
                section_code="CRA",
                section_name="Témata rozvojové spolupráce v gesci ČRA",
                country_name=country_name,
                sector_name=sector_name,
                manager_code="CRA",
                manager_name="Česká rozvojová agentura",
                recipient_key=recipient_key("CRA", recipient_name),
                recipient_name=recipient_name,
                recipient_ico=None,
                project_key=project_key("CRA", "DEVELOPMENT", country_name, recipient_name, project_name),
                project_name=project_name,
                planned_czk=actual_czk,
                actual_czk=actual_czk,
                source_url=CRA_WORKBOOK_URL,
            )
        )

    return rows


def write_snapshot(
    *,
    out_dir: Path,
    snapshot: str,
    rows: list[AidRow],
    metadata_rows: list[dict[str, object]],
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    data_path = out_dir / f"{snapshot}__{OUTPUT_FILE_NAME}"
    fieldnames = [
        "reporting_year",
        "branch_code",
        "branch_name",
        "source_workbook",
        "section_code",
        "section_name",
        "country_name",
        "sector_name",
        "manager_code",
        "manager_name",
        "recipient_key",
        "recipient_name",
        "recipient_ico",
        "project_key",
        "project_name",
        "planned_czk",
        "actual_czk",
        "source_url",
    ]

    with data_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "reporting_year": row.reporting_year,
                    "branch_code": row.branch_code,
                    "branch_name": row.branch_name,
                    "source_workbook": row.source_workbook,
                    "section_code": row.section_code,
                    "section_name": row.section_name,
                    "country_name": row.country_name,
                    "sector_name": row.sector_name,
                    "manager_code": row.manager_code,
                    "manager_name": row.manager_name,
                    "recipient_key": row.recipient_key,
                    "recipient_name": row.recipient_name,
                    "recipient_ico": row.recipient_ico or "",
                    "project_key": row.project_key,
                    "project_name": row.project_name,
                    "planned_czk": f"{row.planned_czk:.2f}",
                    "actual_czk": f"{row.actual_czk:.2f}",
                    "source_url": row.source_url,
                }
            )

    write_sidecar(
        data_path,
        {
            "dataset_code": DATASET_CODE,
            "downloaded_at": datetime.now(UTC).isoformat(),
            "row_count": len(rows),
            "year_count": len(metadata_rows),
            "years": sorted({int(row["year"]) for row in metadata_rows}),
            "sources": metadata_rows,
            "generator": "etl/mzv/fetch_aid_operations.py",
        },
    )
    return data_path


def main() -> None:
    args = parse_args()
    years = sorted(set(args.year))
    unsupported = [year for year in years if year != 2024]
    if unsupported:
        raise SystemExit(f"MZV aid operations currently support only 2024, got: {', '.join(map(str, unsupported))}")

    snapshot = timestamp_label(args.snapshot)
    rows: list[AidRow] = []
    metadata_rows: list[dict[str, object]] = []

    for year in years:
        mzv_payload = fetch_bytes(MZV_WORKBOOK_URL)
        cra_payload = fetch_bytes(CRA_WORKBOOK_URL)
        year_rows = parse_mzv_workbook(mzv_payload, year) + parse_cra_workbook(cra_payload, year)
        rows.extend(year_rows)
        metadata_rows.extend(
            [
                {
                    "year": year,
                    "workbook": "MZV",
                    "source_url": MZV_WORKBOOK_URL,
                    "content_sha256": sha256_bytes(mzv_payload),
                    "row_count": len([row for row in year_rows if row.source_workbook == "MZV"]),
                },
                {
                    "year": year,
                    "workbook": "CRA",
                    "source_url": CRA_WORKBOOK_URL,
                    "content_sha256": sha256_bytes(cra_payload),
                    "row_count": len([row for row in year_rows if row.source_workbook == "CRA"]),
                },
            ]
        )

    rows.sort(key=lambda row: (row.reporting_year, row.branch_code, row.country_name, row.project_name, row.recipient_name))
    data_path = write_snapshot(
        out_dir=args.out_dir,
        snapshot=snapshot,
        rows=rows,
        metadata_rows=metadata_rows,
    )

    print(f"Wrote {data_path}")
    print(f"Years: {', '.join(str(year) for year in years)}")
    print(f"Rows written: {len(rows)}")


if __name__ == "__main__":
    main()
