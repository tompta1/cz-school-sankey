#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

from _common import RAW_ROOT, USER_AGENT, fetch_bytes, sha256_bytes, timestamp_label

DATASET_CODE = "health_zzs_activity_aggregates"
SOURCE_URL = "https://www.nzip.cz/data/vykazy/datove-souhrny/Datovy-souhrn-SSS-07-12-vykaz-a038-zdravotnicka-zachranna-sluzba-2025-01.xlsx"
METHODOLOGY_URL = "https://www.nzip.cz/data/vykazy/metodicke-popisy/Datovy-souhrn-SSS-07-12-vykaz-a038-zdravotnicka-zachranna-sluzba-2025-01.pdf"
OUTPUT_FILE_NAME = "health-zzs-activity-aggregates.csv"
SHEET_NAME = "tabulky absolutni"

INDICATOR_LABELS = {
    "calls_total": "Hovory na tísňových linkách: počet",
    "events_total": "Události řešené na základě tísňové výzvy: počet",
    "patients_total": "Celkový počet pacientů: počet",
    "departures_rlp": "Počet výjezdů ZZS (bez LVS): rychlá lékařská pomoc",
    "departures_rzp": "Počet výjezdů ZZS (bez LVS): rychlá zdravotnická pomoc",
    "departures_rv": "Počet výjezdů ZZS (bez LVS): rendez-vous",
}

INDICATOR_NAMES = {
    "calls_total": "Hovory na tísňových linkách ZZS",
    "events_total": "Události řešené na základě tísňové výzvy",
    "patients_total": "Celkový počet pacientů ZZS",
    "departures_total": "Počet výjezdů ZZS bez LVS",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch annual national ZZS activity aggregates from the official NZIP A038 workbook")
    parser.add_argument("--snapshot", default=None, help="Snapshot label, defaults to YYYYMMDD")
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=RAW_ROOT / DATASET_CODE,
        help="Output directory",
    )
    return parser.parse_args()


def cell_to_number(value: object) -> int:
    if value is None:
        return 0
    return int(round(float(value)))


def parse_rows(xlsx_bytes: bytes) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Missing sheet {SHEET_NAME!r} in A038 workbook")

    sheet = workbook[SHEET_NAME]
    year_row = next(sheet.iter_rows(min_row=7, max_row=7, values_only=True))
    year_columns: dict[int, int] = {}
    for column_index, value in enumerate(year_row):
        if isinstance(value, (int, float)):
            year_columns[column_index] = int(value)

    if not year_columns:
        raise RuntimeError("Could not find annual columns in the A038 workbook")

    values_by_indicator: dict[str, dict[int, int]] = {code: {} for code in INDICATOR_LABELS}
    for row in sheet.iter_rows(values_only=True):
        label = row[0]
        if not isinstance(label, str):
            continue
        normalized_label = label.strip()
        for indicator_code, expected_label in INDICATOR_LABELS.items():
            if normalized_label != expected_label:
                continue
            for column_index, year in year_columns.items():
                values_by_indicator[indicator_code][year] = cell_to_number(row[column_index])

    missing = [code for code, values in values_by_indicator.items() if not values]
    if missing:
        raise RuntimeError(f"Missing indicator rows in A038 workbook: {', '.join(missing)}")

    years = sorted(year_columns.values())
    rows: list[dict[str, object]] = []
    for year in years:
        departures_total = sum(
            values_by_indicator[code][year]
            for code in ("departures_rlp", "departures_rzp", "departures_rv")
        )
        rows.extend(
            [
                {
                    "reporting_year": year,
                    "indicator_code": "patients_total",
                    "indicator_name": INDICATOR_NAMES["patients_total"],
                    "count_value": values_by_indicator["patients_total"][year],
                    "source_url": SOURCE_URL,
                },
                {
                    "reporting_year": year,
                    "indicator_code": "events_total",
                    "indicator_name": INDICATOR_NAMES["events_total"],
                    "count_value": values_by_indicator["events_total"][year],
                    "source_url": SOURCE_URL,
                },
                {
                    "reporting_year": year,
                    "indicator_code": "calls_total",
                    "indicator_name": INDICATOR_NAMES["calls_total"],
                    "count_value": values_by_indicator["calls_total"][year],
                    "source_url": SOURCE_URL,
                },
                {
                    "reporting_year": year,
                    "indicator_code": "departures_total",
                    "indicator_name": INDICATOR_NAMES["departures_total"],
                    "count_value": departures_total,
                    "source_url": SOURCE_URL,
                },
            ]
        )

    return rows


def write_snapshot(*, out_dir: Path, snapshot: str, rows: list[dict[str, object]], metadata: dict[str, object]) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    data_path = out_dir / f"{snapshot}__{OUTPUT_FILE_NAME}"
    fieldnames = [
        "reporting_year",
        "indicator_code",
        "indicator_name",
        "count_value",
        "source_url",
    ]

    with data_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    sidecar_path = data_path.with_suffix(data_path.suffix + ".download.json")
    sidecar_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    return data_path


def main() -> None:
    args = parse_args()
    snapshot = timestamp_label(args.snapshot)
    xlsx_bytes = fetch_bytes(SOURCE_URL)
    rows = parse_rows(xlsx_bytes)

    metadata = {
        "dataset_code": DATASET_CODE,
        "source_url": SOURCE_URL,
        "metadata_url": METHODOLOGY_URL,
        "downloaded_at": datetime.now(UTC).isoformat(),
        "row_count": len(rows),
        "years": sorted({int(row["reporting_year"]) for row in rows}),
        "sha256": sha256_bytes(xlsx_bytes),
        "size_bytes": len(xlsx_bytes),
        "generator": "etl/health/fetch_zzs_activity_aggregates.py",
        "user_agent": USER_AGENT,
    }
    data_path = write_snapshot(out_dir=args.out_dir, snapshot=snapshot, rows=rows, metadata=metadata)
    print(f"Wrote {data_path}")
    print(f"Rows written: {len(rows)}")


if __name__ == "__main__":
    main()
